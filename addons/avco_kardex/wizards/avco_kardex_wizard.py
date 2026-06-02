# -*- coding: utf-8 -*-
import io
import base64
from odoo import models, fields, api
from odoo.exceptions import UserError


class AvcoKardexWizard(models.TransientModel):
    _name = 'avco.kardex.wizard'
    _description = 'Wizard para Kardex AVCO'

    date_from = fields.Date(string='Fecha desde', required=True)
    date_to = fields.Date(
        string='Fecha hasta',
        required=True,
        default=fields.Date.today,
    )
    company_id = fields.Many2one(
        'res.company',
        string='Compañía',
        default=lambda s: s.env.company,
        required=True,
    )
    product_ids = fields.Many2many(
        'product.product',
        string='Productos',
        help='Dejar vacío para incluir todos los productos con costo promedio.',
    )
    line_ids = fields.One2many(
        'avco.kardex.line',
        'wizard_id',
        string='Kardex',
        readonly=True,
    )

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        active_model = self.env.context.get('active_model')
        active_ids = self.env.context.get('active_ids', [])
        if active_model == 'product.product' and active_ids:
            res['product_ids'] = [(6, 0, active_ids)]
        elif active_model == 'product.template' and active_ids:
            tmpls = self.env['product.template'].browse(active_ids)
            res['product_ids'] = [(6, 0, tmpls.mapped('product_variant_ids').ids)]
        return res

    def action_run(self):
        self.ensure_one()
        if self.date_from > self.date_to:
            raise UserError("La fecha desde no puede ser mayor que la fecha hasta.")

        company = self.company_id

        if self.product_ids:
            products = self.product_ids.filtered(
                lambda p: p.categ_id.property_cost_method == 'average'
            )
        else:
            products = self.env['product.product'].search([
                ('categ_id.property_cost_method', '=', 'average'),
            ])

        if not products:
            raise UserError(
                "No hay productos con método de costo promedio para procesar."
            )

        kardex_svc = self.env['avco.kardex']

        # Eliminar líneas anteriores de este wizard
        self.line_ids.unlink()

        vals_list = []
        for product in products:
            rows, totals = kardex_svc.generate_kardex(
                product=product,
                company=company,
                date_from=self.date_from,
                date_to=self.date_to,
            )
            for row in rows:
                vals_list.append({
                    'wizard_id': self.id,
                    'producto': row['producto'],
                    'fecha': row['fecha'],
                    'tipo': row['tipo'],
                    'move_id': row['move_id'] or 0,
                    'ref': row['ref'],
                    'existencia_anterior': row['existencia_anterior'],
                    'entrada': row['entrada'],
                    'salida': row['salida'],
                    'costo_movimiento': row['costo_movimiento'],
                    'nuevo_costo': row['nuevo_costo'],
                    'existencia': row['existencia'],
                    'valor_inventario': row['valor_inventario'],
                })
            vals_list.append({
                'wizard_id': self.id,
                'producto': product.display_name,
                'fecha': str(self.date_to),
                'tipo': 'TOTAL',
                'move_id': 0,
                'ref': '',
                'existencia_anterior': 0.0,
                'entrada': totals['total_in'],
                'salida': totals['total_out'],
                'costo_movimiento': 0.0,
                'nuevo_costo': totals['avco_final'],
                'existencia': totals['qty_final'],
                'valor_inventario': totals['valor_final'],
            })

        if not vals_list:
            raise UserError(
                "No se encontraron movimientos para el período y productos seleccionados."
            )

        self.env['avco.kardex.line'].create(vals_list)

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'avco.kardex.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': self.env.context,
        }

    def action_export_excel(self):
        self.ensure_one()
        if not self.line_ids:
            raise UserError("Genera el kardex primero.")

        rows = [
            {
                'producto': l.producto,
                'fecha': l.fecha,
                'tipo': l.tipo,
                'move_id': l.move_id,
                'ref': l.ref,
                'existencia_anterior': l.existencia_anterior,
                'entrada': l.entrada,
                'salida': l.salida,
                'costo_movimiento': l.costo_movimiento,
                'nuevo_costo': l.nuevo_costo,
                'existencia': l.existencia,
                'valor_inventario': l.valor_inventario,
            }
            for l in self.line_ids
        ]

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise UserError(
                "La librería openpyxl no está disponible. "
                "Instálala con: pip install openpyxl"
            )

        wb = Workbook()
        ws = wb.active
        ws.title = 'Kardex AVCO'

        headers = [
            'Producto', 'Fecha', 'Tipo', 'Move ID', 'Referencia',
            'Exist. Anterior', 'Entrada', 'Salida',
            'Costo Mov.', 'Nuevo Costo', 'Existencia', 'Valor Inventario',
        ]

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E74B5', end_color='2E74B5', fill_type='solid')
        center = Alignment(horizontal='center')

        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        num_fmt = '#,##0.######'
        money_fmt = '#,##0.00'
        num_cols = {6, 7, 8, 9, 10, 11}
        money_col = 12

        tipo_fills = {
            'IN':     PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
            'OUT':    PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
            'INICIO': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
            'INT':    PatternFill(start_color='EDEDED', end_color='EDEDED', fill_type='solid'),
            'TOTAL':  PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
        }
        bold_font = Font(bold=True)

        for row in rows:
            ws.append([
                row['producto'], row['fecha'], row['tipo'], row['move_id'],
                row['ref'], row['existencia_anterior'], row['entrada'],
                row['salida'], row['costo_movimiento'], row['nuevo_costo'],
                row['existencia'], row['valor_inventario'],
            ])
            row_idx = ws.max_row
            fill = tipo_fills.get(row['tipo'])
            is_total = row['tipo'] == 'TOTAL'
            for col_idx, cell in enumerate(ws[row_idx], start=1):
                if fill:
                    cell.fill = fill
                if is_total:
                    cell.font = bold_font
                if col_idx in num_cols:
                    cell.number_format = num_fmt
                elif col_idx == money_col:
                    cell.number_format = money_fmt

        col_widths = [30, 22, 8, 10, 28, 16, 12, 12, 14, 14, 14, 20]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = 'A2'

        buf = io.BytesIO()
        wb.save(buf)
        xlsx_b64 = base64.b64encode(buf.getvalue()).decode()

        filename = f'kardex_{self.date_from}_{self.date_to}.xlsx'

        # Crear attachment temporal y devolver descarga
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'datas': xlsx_b64,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id': self.id,
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }


class AvcoKardexLine(models.TransientModel):
    _name = 'avco.kardex.line'
    _description = 'Línea de Kardex AVCO'
    _order = 'producto asc, fecha asc, id asc'

    wizard_id = fields.Many2one(
        'avco.kardex.wizard',
        required=True,
        ondelete='cascade',
    )
    producto = fields.Char(string='Producto', readonly=True)
    fecha = fields.Char(string='Fecha', readonly=True)
    tipo = fields.Selection(
        [
            ('INICIO', 'Inicio'),
            ('IN', 'Entrada'),
            ('OUT', 'Salida'),
            ('INT', 'Interno'),
            ('TOTAL', 'Total'),
        ],
        string='Tipo',
        readonly=True,
    )
    move_id = fields.Integer(string='Move ID', readonly=True)
    ref = fields.Char(string='Referencia', readonly=True)
    existencia_anterior = fields.Float(
        string='Exist. Anterior',
        digits=(16, 4),
        readonly=True,
    )
    entrada = fields.Float(string='Entrada', digits=(16, 4), readonly=True)
    salida = fields.Float(string='Salida', digits=(16, 4), readonly=True)
    costo_movimiento = fields.Float(
        string='Costo Mov.',
        digits='Product Price',
        readonly=True,
    )
    nuevo_costo = fields.Float(
        string='Nuevo Costo',
        digits='Product Price',
        readonly=True,
    )
    existencia = fields.Float(string='Existencia', digits=(16, 4), readonly=True)
    valor_inventario = fields.Float(
        string='Valor Inventario',
        digits=(16, 2),
        readonly=True,
    )
