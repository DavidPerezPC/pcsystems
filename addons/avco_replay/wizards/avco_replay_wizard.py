# -*- coding: utf-8 -*-
import io
import base64
from odoo import models, fields, api
from odoo.exceptions import UserError


class AvcoReplayWizard(models.TransientModel):
    _name = 'avco.replay.wizard'
    _description = 'Wizard para Replay de AVCO'

    date_from = fields.Date(
        string='Fecha de corte',
        required=True,
        help='Desde esta fecha se reconstruye el AVCO. '
             'La historia anterior se reproduce solo para obtener el AVCO inicial.'
    )
    company_id = fields.Many2one(
        'res.company',
        string='Compañía',
        default=lambda s: s.env.company,
        required=True,
    )
    apply_changes = fields.Boolean(
        string='Aplicar cambios',
        default=False,
        help='Si está desmarcado solo simula y muestra el log (recomendado primero).'
    )
    rewrite_moves = fields.Boolean(
        string='Reescribir valor en stock.move',
        default=False,
        help='Reescribe el campo value de cada stock.move con el valor recalculado. '
             'IRREVERSIBLE. Haz respaldo antes.'
    )
    include_landed_costs = fields.Boolean(
        string='Considerar landed costs',
        default=True,
        help='Suma los ajustes de landed cost al costo unitario de entrada.'
    )
    line_ids = fields.One2many(
        'avco.replay.wizard.line',
        'wizard_id',
        string='Productos',
    )
    result_log = fields.Text(
        string='Log de resultado',
        readonly=True,
    )
    excel_file = fields.Binary(
        string='Archivo Excel',
        readonly=True,
        attachment=False,
    )
    excel_filename = fields.Char(
        string='Nombre archivo',
        readonly=True,
    )

    @api.model
    def default_get(self, fields_list):
        """Al abrir el wizard desde la lista de productos, precarga las líneas."""
        res = super().default_get(fields_list)
        active_model = self.env.context.get('active_model')
        active_ids = self.env.context.get('active_ids', [])

        if active_model == 'product.product' and active_ids:
            products = self.env['product.product'].browse(active_ids).filtered(
                lambda p: p.categ_id.property_cost_method == 'average'
            )
        elif active_model == 'product.template' and active_ids:
            tmpls = self.env['product.template'].browse(active_ids).filtered(
                lambda t: t.categ_id.property_cost_method == 'average'
            )
            products = tmpls.mapped('product_variant_ids')
        else:
            products = self.env['product.product']

        company = self.env.company
        res['line_ids'] = [
            (0, 0, {
                'product_id': p.id,
                'manual_cost': p.with_company(company).standard_price or 0.0,
            })
            for p in products
        ]
        return res

    def action_run(self):
        """Ejecuta el replay para todas las líneas."""
        self.ensure_one()
        if not self.line_ids:
            raise UserError("No hay productos para procesar.")

        replay = self.env['avco.replay']
        all_log = []
        all_rows = []

        for line in self.line_ids:
            all_log.append("")
            all_log.append("=" * 70)
            log, avco, qty, rows = replay.replay_product(
                product=line.product_id,
                company=self.company_id,
                date_from=self.date_from,
                manual_cost=line.manual_cost,
                apply=self.apply_changes,
                rewrite_moves=self.rewrite_moves,
            )
            all_log.extend(log)
            all_rows.extend(rows)
            line.computed_avco = avco
            line.computed_qty = qty

        self.result_log = '\n'.join(all_log)

        if all_rows:
            self.excel_file = self._build_excel(all_rows)
            self.excel_filename = f'avco_replay_{self.date_from}.xlsx'

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'avco.replay.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': self.env.context,
        }

    def _build_excel(self, rows):
        """Genera el archivo Excel con el detalle del replay y retorna base64."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, numbers
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise UserError(
                "La librería openpyxl no está disponible. "
                "Instálala con: pip install openpyxl"
            )

        wb = Workbook()
        ws = wb.active
        ws.title = 'AVCO Replay'

        headers = [
            'Producto', 'Fecha', 'Tipo', 'Move ID', 'Referencia',
            'Existencia Anterior', 'Entrada', 'Salida',
            'Costo Movimiento', 'Nuevo Costo', 'Existencia',
        ]

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E74B5', end_color='2E74B5', fill_type='solid')
        center = Alignment(horizontal='center')

        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        num_fmt = '#,##0.######'
        num_cols = {6, 7, 8, 9, 10, 11}  # columnas numéricas (1-indexed)

        tipo_fills = {
            'IN':     PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
            'OUT':    PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
            'INICIO': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
            'INT':    PatternFill(start_color='EDEDED', end_color='EDEDED', fill_type='solid'),
        }

        for row in rows:
            data = [
                row['producto'],
                row['fecha'],
                row['tipo'],
                row['move_id'],
                row['ref'],
                row['existencia_anterior'],
                row['entrada'],
                row['salida'],
                row['costo_movimiento'],
                row['nuevo_costo'],
                row['existencia'],
            ]
            ws.append(data)
            row_idx = ws.max_row
            fill = tipo_fills.get(row['tipo'])
            for col_idx, cell in enumerate(ws[row_idx], start=1):
                if fill:
                    cell.fill = fill
                if col_idx in num_cols:
                    cell.number_format = num_fmt

        # Ajuste de anchos de columna
        col_widths = [30, 22, 8, 10, 25, 18, 12, 12, 18, 14, 14]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = 'A2'

        buf = io.BytesIO()
        wb.save(buf)
        return base64.b64encode(buf.getvalue())


class AvcoReplayWizardLine(models.TransientModel):
    _name = 'avco.replay.wizard.line'
    _description = 'Línea de producto para replay de AVCO'

    wizard_id = fields.Many2one(
        'avco.replay.wizard',
        required=True,
        ondelete='cascade',
    )
    product_id = fields.Many2one(
        'product.product',
        string='Producto',
        required=True,
    )
    current_standard_price = fields.Float(
        related='product_id.standard_price',
        string='Costo actual',
        readonly=True,
    )
    manual_cost = fields.Float(
        string='Costo para ajustes manuales',
        digits='Product Price',
        required=True,
        help='Costo a usar SOLO cuando un stock.move es un ajuste manual '
             'sin costo propio (sin PO ni factura). Para entradas con fuente '
             'válida se usa esa fuente + landed costs.'
    )
    computed_avco = fields.Float(
        string='AVCO calculado',
        digits='Product Price',
        readonly=True,
    )
    computed_qty = fields.Float(
        string='Qty final',
        readonly=True,
    )
